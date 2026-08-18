from openpyxl import Workbook
from openpyxl.utils import get_column_letter



class SearchExcelWriter:


    def __init__(
            self,
            session,
            logger
    ):

        self.session = session
        self.logger = logger


        self.file = (
            self.session.session_dir
            /
            "search_products.xlsx"
        )




    def save(
            self,
            products
    ):


        try:


            if not products:


                self.logger.warning(
                    "Нет товаров для сохранения"
                )

                return




            self.logger.info(
                "Создание Search Excel: %s",
                self.file
            )



            wb = Workbook()


            ws = wb.active


            ws.title = "Search Products"



            headers = [


                "Дата парсинга",

                "Название",

                "Бренд",

                "Магазин",

                "SKU",

                "Рейтинг карточки",

                "Кол-во отзывов карточки",

                "Старая цена",

                "Цена",

                "Ссылка",

                "Ссылка на магазин"


            ]



            ws.append(
                headers
            )



            for product in products:


                ws.append([


                    product.parse_date,


                    product.title,


                    product.brand,


                    product.supplier,


                    product.sku,


                    product.sku_rating,


                    product.sku_reviews_count,


                    product.price_old,


                    product.price,


                    product.url,


                    product.shop_url


                ])




            self.auto_width(
                ws
            )



            wb.save(
                self.file
            )



            self.logger.info(
                "Search Excel сохранен: %s",
                self.file
            )


            self.logger.info(
                "Количество товаров: %s",
                len(products)
            )



        except Exception as e:


            self.logger.exception(
                "Ошибка SearchExcelWriter: %s",
                e
            )

            raise





    @staticmethod
    def auto_width(
            worksheet
    ):


        for column in worksheet.columns:


            max_length = 0


            column_letter = get_column_letter(
                column[0].column
            )



            for cell in column:


                if cell.value:


                    length = len(
                        str(cell.value)
                    )


                    if length > max_length:

                        max_length = length



            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                60
            )