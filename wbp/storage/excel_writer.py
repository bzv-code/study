from openpyxl import Workbook
from openpyxl.utils import get_column_letter



class ExcelWriter:


    def __init__(
            self,
            session,
            logger
    ):

        self.file = session.excel_file
        self.logger = logger



    def save(
            self,
            products
    ):


        try:


            if not products:

                self.logger.warning(
                    "Нет товаров для сохранения в Excel"
                )

                return



            self.logger.info(
                "Создание Excel файла: %s",
                self.file
            )



            wb = Workbook()


            ws = wb.active


            ws.title = "Products"



            headers = [

                "Дата парсинга",
                "Название",
                "Бренд",
                "Поставщик",
                "SKU",
                "Рейтинг карточки",
                "Кол-во отзывов карточки",
                "Старая цена",
                "Цена",
                "Ссылка"

            ]


            ws.append(
                headers
            )



            for product in products:


                ws.append([

                    product.parse_date,

                    product.title,

                    product.brand,

                    product.supplier,

                    product.sku,

                    product.sku_rating,

                    product.sku_reviews_count,

                    product.price_old,

                    product.price,

                    product.url

                ])




            self.auto_width(
                ws
            )



            wb.save(
                self.file
            )



            self.logger.info(
                "Excel успешно сохранен: %s",
                self.file
            )


            self.logger.info(
                "Количество товаров в Excel: %s",
                len(products)
            )



        except Exception as e:


            self.logger.exception(
                "Ошибка сохранения Excel: %s",
                e
            )

            raise




    @staticmethod
    def auto_width(
            worksheet
    ):


        for column in worksheet.columns:


            max_length = 0


            column_letter = get_column_letter(
                column[0].column
            )


            for cell in column:


                try:


                    if cell.value:

                        length = len(
                            str(cell.value)
                        )


                        if length > max_length:

                            max_length = length



                except Exception:

                    pass



            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                60
            )